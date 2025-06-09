from datetime import datetime
from io import BytesIO

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.hashers import make_password
from django.db.models import Q, Sum
from django.http import HttpResponse, JsonResponse, QueryDict
from django.shortcuts import get_object_or_404, redirect, render
from openpyxl import Workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from accounts.models import Account, Client, Employee
from common.utils import generate_unique_identifier

from .forms import (
    ClientForm,
    EmployeeForm,
    MachineryForm,
    OrderForm,
    ProcessingUnitForm,
    ProductionLogForm,
    WoodInventoryForm,
)
from .models import (
    Machinery,
    Order,
    OrderDetail,
    ProcessingUnit,
    ProductionLog,
    WoodInventory,
)


def custom_404_view(request, exception):
    return render(request, "404.html", status=404)


def home(request):
    # wood_inventories = WoodInventory.objects.all().order_by("-updated_at")
    wood_inventories = WoodInventory.objects.filter(status__iexact="available").order_by(
        "-updated_at"
    )
    context = {
        "wood_inventories": wood_inventories,
    }
    return render(request, "core/clients/pages/home.html", context)


def about(request):
    return render(request, "core/clients/pages/about.html")


def contact(request):
    return render(request, "core/clients/pages/contact.html")


@login_required
def dashboard(request):
    total_users = Account.objects.count()
    total_employees = Employee.objects.count()
    total_clients = Client.objects.count()
    total_wood_inventories = WoodInventory.objects.count()
    total_wood_quantity = WoodInventory.objects.aggregate(Sum("quantity"))["quantity__sum"] or 0
    total_processing_units = ProcessingUnit.objects.count()
    total_production_logs = ProductionLog.objects.count()
    total_orders = Order.objects.count()
    total_revenue = (
        Order.objects.filter(status=Order.OrderStatus.COMPLETED).aggregate(Sum("total_cost"))[
            "total_cost__sum"
        ]
        or 0
    )
    recent_orders = Order.objects.order_by("-created_at")[:5]
    low_stock_woods = WoodInventory.objects.filter(quantity__lt=10).order_by(
        "quantity"
    )  # Example: quantity less than 10

    context = {
        "total_users": total_users,
        "total_employees": total_employees,
        "total_clients": total_clients,
        "total_wood_inventories": total_wood_inventories,
        "total_wood_quantity": total_wood_quantity,
        "total_processing_units": total_processing_units,
        "total_production_logs": total_production_logs,
        "total_orders": total_orders,
        "total_revenue": total_revenue,
        "recent_orders": recent_orders,
        "low_stock_woods": low_stock_woods,
    }
    return render(request, "core/dashboard/pages/dashboard.html", context)


@login_required
def manage_profile(request):
    return render(request, "core/dashboard/pages/manage-profile.html")


@login_required
def manage_users(request):
    query = request.GET.get("search", "")
    users = Account.objects.all()

    if query:
        # Search across multiple fields
        users = users.filter(Q(email__icontains=query))

    if request.GET.get("export") == "excel":
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="users.xlsx"'

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Users"

        headers = [
            "ID",
            "Email",
            "Date Joined",
            "Last Login",
            "Is Active",
            "Is Staff",
        ]
        for col_num, header_title in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num)
            cell.value = header_title

        for row_num, user_obj in enumerate(users, 2):
            sheet.cell(row=row_num, column=1).value = user_obj.id
            sheet.cell(row=row_num, column=2).value = user_obj.email
            sheet.cell(row=row_num, column=3).value = (
                user_obj.date_joined.strftime("%Y-%m-%d %H:%M:%S") if user_obj.date_joined else ""
            )
            sheet.cell(row=row_num, column=4).value = (
                user_obj.last_login.strftime("%Y-%m-%d %H:%M:%S") if user_obj.last_login else ""
            )
            sheet.cell(row=row_num, column=5).value = "Yes" if user_obj.is_active else "No"
            sheet.cell(row=row_num, column=6).value = "Yes" if user_obj.is_staff else "No"
        workbook.save(response)
        return response

    msg = "Are you sure you want to delete this user?"
    context = {
        "users": users,
        "delete_confirm_msg": msg,
    }
    return render(request, "core/dashboard/pages/manage-users.html", context)


@login_required
def delete_user(request, pk):
    user = Account.objects.get(pk=pk)

    if request.method == "POST":
        user.delete()
        return redirect("manage-users")

    return render(request, "core/dashboard/pages/manage-users.html")


@login_required
def manage_employees(request):
    # employee = Employee.objects.get(user=request.user)
    query = request.GET.get("search", "")
    employees = Employee.objects.all()

    if query:
        # Search across multiple fields
        employees = employees.filter(Q(name__icontains=query) | Q(contact_number__icontains=query))

    if request.GET.get("export") == "excel":
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="employees.xlsx"'
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Employees"

        headers = [
            "ID",
            "Name",
            "Email",
            "Contact Number",
            "Date Hired",
            "User Active",
        ]
        for col_num, header_title in enumerate(headers, 1):
            sheet.cell(row=1, column=col_num).value = header_title

        for row_num, emp in enumerate(employees, 2):
            sheet.cell(row=row_num, column=1).value = emp.id
            sheet.cell(row=row_num, column=2).value = emp.name
            sheet.cell(row=row_num, column=3).value = emp.user.email if emp.user else ""
            sheet.cell(row=row_num, column=4).value = emp.contact_number
            sheet.cell(row=row_num, column=5).value = (
                emp.date_hired.strftime("%Y-%m-%d") if emp.date_hired else ""
            )
            sheet.cell(row=row_num, column=6).value = (
                "Yes" if emp.user and emp.user.is_active else "No"
            )

        workbook.save(response)
        return response

    if request.method == "POST":
        form = EmployeeForm(request.POST)
        email = request.POST.get("email")
        password = request.POST.get("password")

        if form.is_valid():
            if email and password is not None:
                hashed_password = make_password(password)
                user = Account.objects.create(email=email, password=hashed_password)
                user.save()

            employee = form.save(commit=False)
            employee.user = user
            employee.save()
            messages.success(request, "Employee created successfully.")
        else:
            print("Form is not valid: ", form.errors)

    context = {
        # "employee": employee,
        "employees": employees,
        "delete_confirm_msg": "Are you sure you want to delete this employee?",
    }
    return render(request, "core/dashboard/pages/manage-employees.html", context)


@login_required
def edit_employee(request, pk):
    # Get the employee to update
    employee = Employee.objects.get(id=pk)
    data = QueryDict(request.body)  # Use QueryDict to parse the request body
    emp_form = EmployeeForm(data, instance=employee)  # Bind the data to the existing instance

    if request.method == "POST":
        if emp_form.is_valid():
            emp_form.save()
            messages.success(request, "Employee updated successfully.")
        else:
            return JsonResponse({"error": emp_form.errors}, status=400)
        return redirect("manage-employees")
    else:
        emp_form = EmployeeForm(data, instance=employee)

    return render(request, "core/dashboard/pages/manage-employees.html")


@login_required
def delete_employee(request, pk):
    employee = Employee.objects.get(pk=pk)

    if request.method == "POST":
        employee.delete()
        return redirect("manage-employees")

    return render(request, "core/dashboard/pages/manage-employees.html")


@login_required
def manage_clients(request):
    clients = Client.objects.all()

    # Get the search query and search_by from the request's GET parameters
    search_query = request.GET.get("search", "")
    search_by = request.GET.get("search_by", "name")  # Default to searching by name

    # Customize the query based on the search_by value
    if search_query and search_by == "name":
        clients = Client.objects.filter(name__icontains=search_query)
    elif search_query and search_by == "contact_number":
        clients = Client.objects.filter(contact_number__contains=search_query)
    else:
        clients = Client.objects.all().order_by("-updated_at")

    if request.GET.get("export") == "excel":
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="clients.xlsx"'
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Clients"

        headers = ["ID", "Name", "Contact Number", "Address", "Registration Date"]
        for col_num, header_title in enumerate(headers, 1):
            sheet.cell(row=1, column=col_num).value = header_title

        for row_num, client_obj in enumerate(clients, 2):
            sheet.cell(row=row_num, column=1).value = client_obj.id
            sheet.cell(row=row_num, column=2).value = client_obj.name
            sheet.cell(row=row_num, column=3).value = client_obj.contact_number
            sheet.cell(row=row_num, column=4).value = client_obj.address
            sheet.cell(row=row_num, column=5).value = (
                client_obj.registration_date.strftime("%Y-%m-%d")
                if client_obj.registration_date
                else ""
            )

        workbook.save(response)
        return response

    if request.method == "POST":
        form = ClientForm(request.POST)
        if form.is_valid():
            form.save(commit=True)
            messages.success(request, "Client created successfully.")
        else:
            print("Form is not valid: ", form.errors)

    context = {
        "clients": clients,
        "delete_confirm_msg": "Are you sure you want to delete this client?",
    }
    return render(request, "core/dashboard/pages/manage-clients.html", context)


@login_required
def edit_client(request, pk):
    # Get the client to update
    client = Client.objects.get(id=pk)
    data = QueryDict(request.body)  # Use QueryDict to parse the request body
    form = ClientForm(data, instance=client)  # Bind the data to the existing instance

    if request.method == "POST":
        if form.is_valid():
            form.save()
            messages.success(request, "Client updated successfully.")
        else:
            return JsonResponse({"error": form.errors}, status=400)
        return redirect("manage-clients")
    else:
        form = ClientForm(data, instance=client)

    return render(request, "core/dashboard/pages/manage-clients.html")


@login_required
def delete_client(request, pk):
    client = Client.objects.get(pk=pk)

    if request.method == "POST":
        client.delete()
        return redirect("manage-clients")

    return render(request, "core/dashboard/pages/manage-clients.html")


@login_required
def manage_wood_inventories(request):
    wood_inventories = WoodInventory.objects.all().order_by("-updated_at")

    # Get the search query and search_by from the request's GET parameters
    search_query = request.GET.get("search", "")

    if search_query:
        wood_inventories = wood_inventories.filter(
            Q(wood_type__icontains=search_query) | Q(grade__contains=search_query)
        )

    if request.GET.get("export") == "excel":
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="wood_inventories.xlsx"'
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Wood Inventories"

        headers = [
            "ID",
            "Wood Type",
            "Grade",
            "Quantity (m³)",
            "Unit Price (LAK)",
            "Source",
            "Arrival Date",
            "Status",
            "Image URL",
        ]
        for col_num, header_title in enumerate(headers, 1):
            sheet.cell(row=1, column=col_num).value = header_title

        for row_num, wood in enumerate(wood_inventories, 2):
            sheet.cell(row=row_num, column=1).value = wood.id
            sheet.cell(row=row_num, column=2).value = wood.wood_type
            sheet.cell(row=row_num, column=3).value = wood.grade
            sheet.cell(row=row_num, column=4).value = wood.quantity
            sheet.cell(row=row_num, column=5).value = wood.unit_price
            sheet.cell(row=row_num, column=6).value = wood.source
            sheet.cell(row=row_num, column=7).value = (
                wood.arrival_date.strftime("%Y-%m-%d") if wood.arrival_date else ""
            )
            sheet.cell(row=row_num, column=8).value = wood.status
            sheet.cell(row=row_num, column=9).value = (
                request.build_absolute_uri(wood.image.url) if wood.image else ""
            )
        workbook.save(response)
        return response

    if request.method == "POST":
        form = WoodInventoryForm(request.POST, request.FILES)
        if form.is_valid():
            form.save(commit=True)
            # messages.success(request, "Wood inventory created successfully.")
        else:
            print("Form is not valid: ", form.errors)

    context = {
        "wood_inventories": wood_inventories,
        "delete_confirm_msg": "Are you sure you want to delete this wood inventory?",
    }
    return render(request, "core/dashboard/pages/manage-wood-inventories.html", context)


@login_required
def edit_wood_inventory(request, pk):
    wood_inventory = WoodInventory.objects.get(id=pk)

    if request.method == "POST":
        form = WoodInventoryForm(request.POST, request.FILES, instance=wood_inventory)

        if form.is_valid():
            form.save()
            messages.success(request, "Wood inventory updated successfully.")
        else:
            print("Error: ", form.errors)
        return redirect("manage-wood-inventories")
    else:
        form = WoodInventoryForm(instance=wood_inventory)

    return render(request, "core/dashboard/pages/manage-wood-inventories.html")


@login_required
def delete_wood_inventory(request, pk):
    wood_inventory = WoodInventory.objects.get(pk=pk)

    if request.method == "POST":
        wood_inventory.delete()
        return redirect("manage-wood-inventories")

    return render(request, "core/dashboard/pages/manage-wood-inventories.html")


@login_required
def manage_processing_units(request):
    processing_units = ProcessingUnit.objects.all().order_by("-updated_at")

    # Get the search query and search_by from the request's GET parameters
    search_query = request.GET.get("search", "")

    if search_query:
        processing_units = processing_units.filter(Q(name__icontains=search_query))

    if request.GET.get("export") == "excel":
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="processing_units.xlsx"'
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Processing Units"

        headers = ["ID", "Name", "Description", "Capacity", "Status"]
        for col_num, header_title in enumerate(headers, 1):
            sheet.cell(row=1, column=col_num).value = header_title

        for row_num, unit in enumerate(processing_units, 2):
            sheet.cell(row=row_num, column=1).value = unit.id
            sheet.cell(row=row_num, column=2).value = unit.name
            sheet.cell(row=row_num, column=3).value = unit.description
            sheet.cell(row=row_num, column=4).value = unit.capacity
            sheet.cell(row=row_num, column=5).value = unit.status

        workbook.save(response)
        return response

    if request.method == "POST":
        form = ProcessingUnitForm(request.POST)
        if form.is_valid():
            form.save(commit=True)
            # messages.success(request, "Processing unit created successfully.")
        else:
            print("Form is not valid: ", form.errors)

    context = {
        "processing_units": processing_units,
        "delete_confirm_msg": "Are you sure you want to delete this unit?",
    }
    return render(request, "core/dashboard/pages/manage-processing-units.html", context)


@login_required
def edit_processing_unit(request, pk):
    # Get the processing_unit to update
    processing_unit = ProcessingUnit.objects.get(id=pk)
    data = QueryDict(request.body)  # Use QueryDict to parse the request body
    form = ProcessingUnitForm(
        data, instance=processing_unit
    )  # Bind the data to the existing instance

    if request.method == "POST":
        if form.is_valid():
            form.save()
            # messages.success(request, "Data updated successfully.")
        else:
            return JsonResponse({"error": form.errors}, status=400)
        return redirect("manage-processing-units")
    else:
        form = ProcessingUnitForm(data, instance=processing_unit)

    return render(request, "core/dashboard/pages/manage-processing-units.html")


@login_required
def delete_processing_unit(request, pk):
    processing_unit = ProcessingUnit.objects.get(pk=pk)

    if request.method == "POST":
        processing_unit.delete()
        return redirect("manage-processing-units")

    return render(request, "core/dashboard/pages/manage-processing-units.html")


@login_required
def manage_production_logs(request):
    production_logs = ProductionLog.objects.all().order_by("-updated_at")
    wood_inventories = WoodInventory.objects.all()
    processing_units = ProcessingUnit.objects.all()
    employees = Employee.objects.all()

    # Get the search query and search_by from the request's GET parameters
    search_query = request.GET.get("search", "")

    if search_query:
        production_logs = production_logs.filter(Q(status__icontains=search_query))

    if request.GET.get("export") == "excel":
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="production_logs.xlsx"'
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Production Logs"

        headers = [
            "ID",
            "Wood Inventory",
            "Quantity Processed",
            "Processing Unit",
            "Employee",
            "Processing Date",
            "Status",
            "Remarks",
        ]
        for col_num, header_title in enumerate(headers, 1):
            sheet.cell(row=1, column=col_num).value = header_title

        for row_num, log in enumerate(production_logs, 2):
            sheet.cell(row=row_num, column=1).value = log.id
            sheet.cell(row=row_num, column=2).value = str(log.wood_inventory)
            sheet.cell(row=row_num, column=3).value = log.quantity_processed
            sheet.cell(row=row_num, column=4).value = str(log.processing_unit)
            sheet.cell(row=row_num, column=5).value = str(log.employee) if log.employee else ""
            sheet.cell(row=row_num, column=6).value = (
                log.processing_date.strftime("%Y-%m-%d") if log.processing_date else ""
            )
            sheet.cell(row=row_num, column=7).value = log.status
            sheet.cell(row=row_num, column=8).value = log.remarks

        workbook.save(response)
        return response

    if request.method == "POST":
        form = ProductionLogForm(request.POST)
        if form.is_valid():
            form.save(commit=True)
            # messages.success(request, "Processing unit created successfully.")
        else:
            print("Form is not valid: ", form.errors)

    context = {
        "employees": employees,
        "production_logs": production_logs,
        "wood_inventories": wood_inventories,
        "processing_units": processing_units,
        "delete_confirm_msg": "Are you sure you want to delete this production log?",
    }
    return render(request, "core/dashboard/pages/manage-production-logs.html", context)


@login_required
def edit_production_log(request, pk):
    # Get the production_log to update
    production_log = ProductionLog.objects.get(id=pk)
    data = QueryDict(request.body)  # Use QueryDict to parse the request body
    form = ProductionLogForm(
        data, instance=production_log
    )  # Bind the data to the existing instance

    if request.method == "POST":
        if form.is_valid():
            form.save()
            # messages.success(request, "Data updated successfully.")
        else:
            return JsonResponse({"error": form.errors}, status=400)
        return redirect("manage-production-logs")
    else:
        form = ProductionLogForm(data, instance=production_log)

    return render(request, "core/dashboard/pages/manage-production-logs.html")


@login_required
def delete_production_log(request, pk):
    production_log = ProductionLog.objects.get(pk=pk)

    if request.method == "POST":
        production_log.delete()
        return redirect("manage-production-logs")

    return render(request, "core/dashboard/pages/manage-production-logs.html")


@login_required
def manage_machineries(request):
    machineries = Machinery.objects.all().order_by("-updated_at")
    processing_units = ProcessingUnit.objects.all().order_by("-created_at")

    # Get the search query and search_by from the request's GET parameters
    search_query = request.GET.get("search", "")

    if request.GET.get("export") == "excel":
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="machineries.xlsx"'
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Machineries"

        headers = [
            "ID",
            "Name",
            "Processing Unit",
            "Description",
            "Status",
            "Purchase Date",
            "Last Maintenance Date",
        ]
        for col_num, header_title in enumerate(headers, 1):
            sheet.cell(row=1, column=col_num).value = header_title

        for row_num, machine in enumerate(machineries, 2):
            sheet.cell(row=row_num, column=1).value = machine.id
            sheet.cell(row=row_num, column=2).value = machine.name
            sheet.cell(row=row_num, column=3).value = str(machine.processing_unit)
            sheet.cell(row=row_num, column=4).value = machine.description
            sheet.cell(row=row_num, column=5).value = machine.status
            sheet.cell(row=row_num, column=6).value = (
                machine.purchase_date.strftime("%Y-%m-%d") if machine.purchase_date else ""
            )
            sheet.cell(row=row_num, column=7).value = (
                machine.last_maintenance_date.strftime("%Y-%m-%d")
                if machine.last_maintenance_date
                else ""
            )

        workbook.save(response)
        return response

    if search_query:
        machineries = machineries.filter(
            Q(name__icontains=search_query) | Q(status__contains=search_query)
        )

    if request.method == "POST":
        form = MachineryForm(request.POST)
        if form.is_valid():
            form.save(commit=True)
            # messages.success(request, "Processing unit created successfully.")
        else:
            print("Form is not valid: ", form.errors)

    context = {
        "machineries": machineries,
        "processing_units": processing_units,
        "date1": 3,
        "date2": 4,
        "delete_confirm_msg": "Are you sure you want to delete this machinery?",
    }
    return render(request, "core/dashboard/pages/manage-machineries.html", context)


@login_required
def edit_machinery(request, pk):
    # Get the machinery to update
    machinery = Machinery.objects.get(id=pk)
    data = QueryDict(request.body)  # Use QueryDict to parse the request body
    form = MachineryForm(data, instance=machinery)  # Bind the data to the existing instance

    if request.method == "POST":
        if form.is_valid():
            form.save()
            # messages.success(request, "Data updated successfully.")
        else:
            return JsonResponse({"error": form.errors}, status=400)
        return redirect("manage-machineries")
    else:
        form = MachineryForm(data, instance=machinery)

    return render(request, "core/dashboard/pages/manage-machineries.html")


@login_required
def delete_machinery(request, pk):
    machinery = Machinery.objects.get(pk=pk)

    if request.method == "POST":
        machinery.delete()
        return redirect("manage-machineries")

    return render(request, "core/dashboard/pages/manage-machineries.html")


@login_required
def manage_orders(request):
    orders = Order.objects.all().order_by("-updated_at")
    clients = Client.objects.all().order_by("-updated_at")
    wood_inventories = WoodInventory.objects.all().order_by("-updated_at")

    # Get the search query and search_by from the request's GET parameters
    search_query = request.GET.get("search", "")

    if request.GET.get("export") == "excel":
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="orders.xlsx"'
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Orders"

        headers = [
            "ID",
            "Order Number",
            "Client",
            "User",
            "Order Date",
            "Status",
            "Payment Method",
            "Total Cost (LAK)",
            "Description",
        ]
        for col_num, header_title in enumerate(headers, 1):
            sheet.cell(row=1, column=col_num).value = header_title

        for row_num, order_obj in enumerate(orders, 2):
            sheet.cell(row=row_num, column=1).value = order_obj.id
            sheet.cell(row=row_num, column=2).value = order_obj.order_number
            sheet.cell(row=row_num, column=3).value = (
                str(order_obj.client) if order_obj.client else ""
            )
            sheet.cell(row=row_num, column=4).value = str(order_obj.user) if order_obj.user else ""
            sheet.cell(row=row_num, column=5).value = (
                order_obj.order_date.strftime("%Y-%m-%d %H:%M:%S") if order_obj.order_date else ""
            )
            sheet.cell(row=row_num, column=6).value = order_obj.status
            sheet.cell(row=row_num, column=7).value = order_obj.payment_method
            sheet.cell(row=row_num, column=8).value = order_obj.total_cost
            sheet.cell(row=row_num, column=9).value = order_obj.description

        workbook.save(response)
        return response

    if search_query:
        orders = orders.filter(
            Q(order_number__icontains=search_query) | Q(status__contains=search_query)
        )

    if request.method == "POST":
        form = OrderForm(request.POST)
        if form.is_valid():
            order = form.save(commit=False)
            order.order_number = (
                f'ORD-{datetime.now().strftime("%Y%m%d")}-{generate_unique_identifier()}'
            )
            order.user = request.user
            order.status = "completed"
            order.save()

            # Handle OrderDetails
            wood_inventory_ids = request.POST.getlist("wood_inventory")
            quantities = request.POST.getlist("qty")
            prices = request.POST.getlist("price_per_unit")

            total_cost = 0
            for wood_id, qty, price in zip(wood_inventory_ids, quantities, prices):
                wood = WoodInventory.objects.get(id=wood_id)
                total_price = float(qty) * float(price)
                OrderDetail.objects.create(
                    order=order,
                    wood_inventory=wood,
                    qty=qty,
                    price_per_unit=price,
                    total_price=total_price,
                )
                total_cost += total_price

            order.total_cost = total_cost
            order.save()

            messages.success(request, "Order created successfully.")
            return redirect("manage-orders")
        else:
            # messages.error(request, "Error creating order. Please check the form.")
            print("Form is not valid: ", form.errors)

    context = {
        "orders": orders,
        "clients": clients,
        "wood_inventories": wood_inventories,
        "delete_confirm_msg": "Are you sure you want to delete this order?",
    }
    return render(request, "core/dashboard/pages/manage-orders.html", context)


@login_required
def view_order(request, pk):
    """
    Displays the details of a specific order.

    Args:
        request: The HTTP request object.
        pk: The primary key of the order to view.

    Returns:
        An HTTP response rendering the order details template.
    """
    order = get_object_or_404(Order, pk=pk)
    order_details = order.orderdetail_set.all()

    context = {
        "order": order,
        "order_details": order_details,
    }
    return render(request, "core/dashboard/pages/view-order.html", context)


@login_required
def edit_order(request, pk):
    order = get_object_or_404(Order, pk=pk)
    clients = Client.objects.all().order_by("-updated_at")
    wood_inventories = WoodInventory.objects.all().order_by("-updated_at")

    if request.method == "POST":
        form = OrderForm(request.POST, instance=order)
        if form.is_valid():
            order = form.save()
            # Handle existing order details
            existing_detail_ids = []
            for detail in order.orderdetail_set.all():
                detail_id = str(detail.id)
                existing_detail_ids.append(detail_id)
                wood_inventory_id = request.POST.get(f"wood_inventory_{detail_id}")
                qty = request.POST.get(f"qty_{detail_id}")
                price_per_unit = request.POST.get(f"price_per_unit_{detail_id}")

                if wood_inventory_id and qty and price_per_unit:
                    wood = get_object_or_404(WoodInventory, id=wood_inventory_id)
                    detail.wood_inventory = wood
                    detail.qty = qty
                    detail.price_per_unit = price_per_unit
                    detail.total_price = float(qty) * float(price_per_unit)
                    detail.save()

            # Handle new order details
            new_wood_inventory_ids = request.POST.getlist("wood_inventory")
            new_quantities = request.POST.getlist("qty")
            new_prices = request.POST.getlist("price_per_unit")

            for wood_id, qty, price in zip(new_wood_inventory_ids, new_quantities, new_prices):
                if wood_id and qty and price:
                    wood = get_object_or_404(WoodInventory, id=wood_id)
                    OrderDetail.objects.create(
                        order=order,
                        wood_inventory=wood,
                        qty=qty,
                        price_per_unit=price,
                        total_price=float(qty) * float(price),
                    )

            # Handle deleted order details
            for detail_id in existing_detail_ids:
                if f"wood_inventory_{detail_id}" not in request.POST:
                    detail = get_object_or_404(OrderDetail, id=detail_id)
                    detail.delete()

            # Recalculate total cost
            total_cost = sum(detail.total_price for detail in order.orderdetail_set.all())
            order.total_cost = total_cost
            order.save()

            messages.success(request, "Order updated successfully.")
            return redirect("manage-orders")
        else:
            messages.error(request, "Error updating order. Please check the form.")
            print("Form is not valid: ", form.errors)
    else:
        form = OrderForm(instance=order)

    context = {
        "form": form,
        "order": order,
        "clients": clients,
        "wood_inventories": wood_inventories,
    }
    return render(request, "core/dashboard/pages/manage-orders.html", context)


@login_required
def delete_order(request, pk):
    order = get_object_or_404(Order, pk=pk)

    if request.method == "POST":
        order.delete()
        messages.success(request, "Order deleted successfully.")
        return redirect("manage-orders")

    return render(request, "core/dashboard/pages/manage-orders.html")


@login_required
def manage_products(request):
    context = {
        "delete_confirm_msg": "Are you sure you want to delete this product?",
    }
    return render(request, "core/dashboard/pages/manage-products.html", context)


# PDF generation
def generate_order_pdf(request, order_id):
    order = get_object_or_404(Order, id=order_id)
    order_details = OrderDetail.objects.filter(order=order)

    # Create a file-like buffer to receive PDF data
    buffer = BytesIO()

    # Create the PDF object, using the buffer as its "file."
    doc = SimpleDocTemplate(
        buffer, pagesize=letter, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18
    )

    # Container for the 'Flowable' objects
    elements = []

    # A large collection of style sheets pre-made for us
    styles = getSampleStyleSheet()

    # Document title
    elements.append(Paragraph(f"Order #{order.order_number}", styles["Title"]))
    elements.append(Spacer(1, 12))

    # Order information
    elements.append(Paragraph("Order Information", styles["Heading2"]))
    elements.append(Paragraph(f"Client: {order.client.name}", styles["Normal"]))
    elements.append(Paragraph(f"Payment Method: {order.payment_method}", styles["Normal"]))
    elements.append(Paragraph(f"Status: {order.status}", styles["Normal"]))
    elements.append(Paragraph(f"Total Cost: ${order.total_cost}", styles["Normal"]))
    elements.append(
        Paragraph(f"Created At: {order.created_at.strftime('%B %d, %Y %H:%M')}", styles["Normal"])
    )
    elements.append(
        Paragraph(f"Updated At: {order.updated_at.strftime('%B %d, %Y %H:%M')}", styles["Normal"])
    )
    elements.append(Spacer(1, 12))

    if order.description:
        elements.append(Paragraph("Description", styles["Heading2"]))
        elements.append(Paragraph(order.description, styles["Normal"]))
        elements.append(Spacer(1, 12))

    # Order items
    elements.append(Paragraph("Order Items", styles["Heading2"]))

    # Create table for order items
    data = [["#", "Wood Inventory", "Quantity", "Price/Unit", "Total Price"]]
    for index, detail in enumerate(order_details, start=1):
        data.append(
            [
                str(index),
                f"{detail.wood_inventory.wood_type} - {detail.wood_inventory.grade}",
                str(detail.qty),
                f"${detail.price_per_unit}",
                f"${detail.total_price}",
            ]
        )

    table = Table(data)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 14),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                ("TEXTCOLOR", (0, 1), (-1, -1), colors.black),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
                ("FONTSIZE", (0, 1), (-1, -1), 12),
                ("TOPPADDING", (0, 1), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 1), (-1, -1), 6),
                ("GRID", (0, 0), (-1, -1), 1, colors.black),
            ]
        )
    )

    elements.append(table)

    # Build the PDF
    doc.build(elements)

    # Get the value of the BytesIO buffer and write it to the response
    pdf = buffer.getvalue()
    buffer.close()
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="order_{order.order_number}.pdf"'
    response.write(pdf)

    return response
